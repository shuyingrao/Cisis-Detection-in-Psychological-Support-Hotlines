from pathlib import Path
import pandas as pd
import numpy as np
import json
import os
from scipy.stats import mode
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


base_dir = ".../hotline"
label_name = ['情绪状态','自杀意念','自杀计划', '高危']
test_label_path = Path(base_dir) / "data" / "labels" / "2023_Y.csv"
label_data = pd.read_csv(test_label_path)
y_test = label_data[label_name].values


prediction_dir = Path(base_dir) / "outputs" / "predictions" / "text" / "GPT_CNN.npy"

y_preds = np.load(prediction_dir)
# 假设我们有实际标签y_true和保存的预测概率结果y_preds
n_repeats, n_samples, n_labels = y_preds.shape

predictions = np.zeros((n_repeats, n_samples, n_labels), dtype=int)

for i in range(n_repeats):
    for label in range(n_labels):
        predictions[i, :, label] = (y_preds[i, :, label] >= 0.5).astype(int)

final_predictions = np.zeros((n_samples, n_labels), dtype=int)

# Compute the majority vote for each sample and each label
for label in range(n_labels):
    final_predictions[:, label] = mode(predictions[:, :, label], axis=0).mode.flatten()

# %%
# 高危标签的索引
y_true = y_test[:, 3]
y_pred = final_predictions[:, 3]

# 初始化存储TP, FP, TN, FN的索引列表
tp_indices = np.where((y_true == 1) & (y_pred == 1))[0]  # true positive
fp_indices = np.where((y_true == 0) & (y_pred == 1))[0]  # false positive
tn_indices = np.where((y_true == 0) & (y_pred == 0))[0]  # true negative
fn_indices = np.where((y_true == 1) & (y_pred == 0))[0]  # false negative

sentences_dir  = Path(base_dir) / "data" / "Sentences" / "2023_Y"
files = os.listdir(sentences_dir)
subject = files[0]
test_example = pd.read_excel(os.path.join(sentences_dir, subject))
text = str(test_example['content'].tolist())

# %%
# tp_indices, fp_indices, tn_indices, fn_indices
rows = []
true_labels = []
output_file = Path(base_dir) / "data" / "explanation" / "true_positive.xlsx"
for idx in tp_indices:
    subject = files[idx]
    test_example = pd.read_excel(os.path.join(sentences_dir, subject))
    text = str(test_example['content'].tolist())
    predicted_labels = final_predictions[idx]
    true_labels.append(y_test[idx])
    rows.append([subject] + [text] + predicted_labels.tolist())

# 创建DataFrame
columns = ['Subject','Text'] + label_name
df = pd.DataFrame(rows, columns=columns)

# 保存到Excel文件
df.to_excel(output_file, index=False)

# 打开Excel文件并应用条件格式
workbook = load_workbook(output_file)
worksheet = workbook.active

# 应用条件格式：如果预测标签与真实标签不一致，使用红色填充
fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

for i in range(len(df)):
    for j in range(n_labels):
        true_value = true_labels[i][j]
        predicted_value = df.iloc[i][j + 2]  # 第 j+1 列是预测标签
        if true_value != predicted_value:
            worksheet.cell(row=i + 2, column=j + 3).fill = fill

# 保存带条件格式的文件
workbook.save(output_file)

## 
import pandas as pd

output_file = Path(base_dir) / "data" / "explanation" / "true_positive.xlsx"

data = pd.read_excel(output_file)
result_str = []
for index, row in data.iterrows():
    # 提取出文本内容，假设内容在 'Text' 列
    sentences = row['Text']

    # 将标签转换为对应的文本格式
    emotion_state = "抑郁" if row['情绪状态'] else "非抑郁"
    suicide_ideation = "有" if row['自杀意念'] else "无"
    suicide_plan = "有" if row['自杀计划'] else "无"
    high_risk = "高危" if row['高危'] else "非高危"

    # 构建标签字符串
    label_str = f'{{"情绪状态": "{emotion_state}", "自杀意念": "{suicide_ideation}", "自杀计划": "{suicide_plan}", "是否高危": "{high_risk}"}}'

    # 生成所需的格式化字符串
    result_str.append(f'考虑下面的一则心理热线来电的通话内容，以句子列表的形式提供。已知对于该来电的情绪状态和自杀危险性的标签判断为：{label_str}，请你分析来电内容，对来电的标签进行解释，在解释时注明原文依据。来电内容如下：{sentences}')